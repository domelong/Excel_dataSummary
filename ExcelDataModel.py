import openpyxl as opx
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
import os, re
import time

class ExcelDataModel:
    def __init__(self, workbook=opx.Workbook()) -> None:
        """
        ExcelDataModel 数据模型对象 数据获取 数据分析 数据操作
        """
        self.workbook = workbook
        self.fileExtension = (".xlsx", ".xlrd")
        
    # 检查文件拓展名
    def __checkFileExtension(self, path):
        if len(path) >= 5: fileExtension = path[-5:]
        if fileExtension not in self.fileExtension: raise Exception(f"{path}该文件不是xlsx或xlrd文件")
        return True
    
    # sheet 合并单元格复制
    def __copyMergedCells(self, old_sheetObj, new_sheetObj):
        try:
            old_sheetObj_merged_cells = old_sheetObj.merged_cells.ranges
            if old_sheetObj_merged_cells:
                for merged_cell in old_sheetObj_merged_cells:
                    new_sheetObj.merge_cells(str(merged_cell))
        except:
            return -1

    # sheet 数据复制
    def __copyData(self, old_sheetObj, new_sheetObj):
        for row in old_sheetObj.iter_rows():
            for cell in row:
                value = cell.value
                rowindex = cell.row
                colindex = cell.column
                new_sheetObj.cell(rowindex, colindex, value)

    # sheet style复制
    def __copyStyle(self, old_sheetObj, new_sheetObj):
        for row_index in range(1, old_sheetObj.max_row + 1):
            # 遍历旧工作表的列
            for col_index in range(1, old_sheetObj.max_column + 1):
                old_cell = old_sheetObj.cell(row=row_index, column=col_index)
                new_cell = new_sheetObj.cell(row=row_index, column=col_index)
                # 复制字体样式
                if old_cell.font:
                    new_cell.font = Font(name=old_cell.font.name,
                                    size=old_cell.font.size,
                                    bold=old_cell.font.bold,
                                    italic=old_cell.font.italic,
                                    color=old_cell.font.color)
                # 复制填充样式 仅纯色填充
                if old_cell.fill:
                    new_cell.fill = PatternFill(fill_type=old_cell.fill.fill_type,
                                    start_color=old_cell.fill.start_color,
                                    end_color=old_cell.fill.end_color)
                # 复制边框样式
                if old_cell.border:
                    new_cell.border = Border(left=Side(style=old_cell.border.left.style,
                                                color=old_cell.border.left.color),
                                        right=Side(style=old_cell.border.right.style,
                                                    color=old_cell.border.right.color),
                                        top=Side(style=old_cell.border.top.style,
                                                color=old_cell.border.top.color),
                                        bottom=Side(style=old_cell.border.bottom.style,
                                                    color=old_cell.border.bottom.color))
                # 复制对齐方式
                if old_cell.alignment:
                    new_cell.alignment = Alignment(horizontal=old_cell.alignment.horizontal,
                                            vertical=old_cell.alignment.vertical,
                                            text_rotation=old_cell.alignment.text_rotation,
                                            wrap_text=old_cell.alignment.wrap_text)

    # sheet复制
    def __copySheet(self, old_sheetObj, new_sheetObj, keep_style=True):
            if not isinstance(old_sheetObj, opx.worksheet.worksheet.Worksheet): 
                raise Exception("old_sheetObj应该是一个Worksheet对象")
            if not isinstance(new_sheetObj, opx.worksheet.worksheet.Worksheet): 
                raise Exception("new_sheetObj应该是一个Worksheet对象")
            
            min_row = old_sheetObj.min_row
            max_row = old_sheetObj.max_row
            old_range = old_sheetObj[min_row:max_row]
            new_range = new_sheetObj[min_row:max_row]
            # 1. 数据先加进去
            self.__copyData(old_sheetObj, new_sheetObj)

            if keep_style:
                # 2. 处理合并单元格
                self.__copyMergedCells(old_sheetObj, new_sheetObj)
                        
                # 3. 处理样式
                self.__copyStyle(old_sheetObj, new_sheetObj)

    # sheet合并
    def __mergeSheet(self, sheetObj, target_sheetObj):
        for row_range in sheetObj.iter_rows():
            row_list = [cell.value for cell in row_range]
            target_sheetObj.append(row_list)

        self.__copyStyle(sheetObj, target_sheetObj)
        self.__copyMergedCells(sheetObj, target_sheetObj)

    # 创建文件夹将多文件保存到本地 -> workbook_list
    def __save_files_in_folder(self, wb_path_dict, folderNameRemark)-> list:
        """
        params:
        wb_path_dict: dict {workbookName: opx.Workbook}
        """
        # 时间戳
        timestamp = time.strftime("%Y_%m_%d_%H%M%S", time.localtime())
        newfolder = f".\\newfiles_{folderNameRemark}_" + timestamp
        os.mkdir(newfolder)
        wblist = []
        for wbname, wb in wb_path_dict.items(): 
            if re.search(r"\\", wbname): raise Exception(r"文件名中不该有\\")
            wb.save(f".\{newfolder}\{wbname}.xlsx")
            wblist.append(wb)
        return wblist
    
    # 将单文件保存在本地
    def __save_file(self, workbook, filename, NameRemark):
        if not isinstance(workbook, opx.Workbook): raise Exception("workbook必须是一个openpyxl.Workbook对象")
        # 时间戳
        timestamp = time.strftime("%Y_%m_%d_%H%M%S", time.localtime())
        workbook.save(f".\\{filename}_{NameRemark}_{timestamp}.xlsx")

    # 工作簿内多工作表拆分成多工作簿
    def __splitSheets_to_workbooks(self, path, sheets=[])-> dict:
        """
        params:
        path: str 目标文件路径
        sheets: list 工作表名称列表
        """
        if not os.path.isfile(path): raise Exception("路径不是个文件")
        self.__checkFileExtension(path)
        self.workbook = opx.load_workbook(path)
        wb = self.workbook
        worksheets = wb.worksheets
        if sheets:
            if len(sheets) == 2:
                if isinstance(sheets[0], int) and isinstance(sheets[1], int):
                    startIndex = sheets[0] - 1
                    endIndex = sheets[1]
                    worksheets = wb.worksheets[startIndex, endIndex]
            elif isinstance(sheets[0], str) and isinstance(sheets[1], str):
                worksheets = [wb[sheetname] for sheetname in sheets] if sheets else wb.worksheets

        filename = os.path.basename(path)
        count = 1
        new_wbdict = {}
        for sheet in worksheets:
            new_wb = opx.Workbook()
            new_sheetname = sheet.title
            new_sheet = new_wb.active
            new_sheet.title = new_sheetname
            self.__copySheet(sheet, new_sheet)
            new_filename =  f"{count}_" + new_sheetname + filename
            new_wbdict[new_filename] = new_wb
            count += 1
        return new_wbdict

    # 获取单工作簿多工作表合并后数据输出为workbook对象
    def __getdata_to_workbook(self, path, area={}, sheets=[]) -> opx.Workbook:
        """
        params:
        path: str 文件夹路径
        area: dict 一维字典
        sheets: list 工作表名称列表
        """
        if not os.path.isfile(path): raise Exception(f"{path}不是一个文件")
        if not isinstance(area, dict): raise Exception("area参数类型应为一维dict对象")
        if not isinstance(sheets, list): raise Exception("sheets参数类型应为一维list对象")
        self.__checkFileExtension(path)
        self.workbook = opx.load_workbook(path, read_only=True, data_only=True)
        wb = self.workbook
        wslist = [wb[sheetname] for sheetname in sheets] if sheets else wb.worksheets
        new_rowlist = []
        new_wb = opx.Workbook()
        new_ws = new_wb.active
        # 有area
        if area:
            for ws in wslist:
                rowlist = []
                for field, area_map in area.items():
                    ws_range = ws[area_map]
                    rowcount = 1
                    for row in ws_range:
                        cellcount = 1
                        for cell in row:
                            field_index = str(field) + f"_{rowcount}_{cellcount}"
                            rowlist.append(field_index)
                            rowlist.append(cell.value)
                            cellcount += 1
                        rowcount += 1
                new_rowlist.append(rowlist)
                new_ws.append(rowlist)
        else:
            # 无area
            for ws in wslist:
                for row in ws.iter_rows():
                    rowlist = []
                    for cell in row:
                        if cell.value:
                            rowlist.append(cell.value)
                    if len(rowlist) <= 1:continue
                    new_rowlist.append(rowlist)
                    new_ws.append(rowlist)
        return new_wb
        
    # 获取多工作簿多工作表数据输出为workbook_map_dict 后续调整异步
    def __getdata_for_workbooks(self, path, area={}, filenames=[], sheets=[]) -> dict: 
        """
        params:
        path: str 文件夹路径
        area: dict 一维字典
        filename: list 工作簿名称列表
        sheets: list 工作表名称列表
        """
        if not isinstance(filenames, list): raise Exception("filenames参数类型应为一维list对象")
        if os.path.isfile(path): raise Exception("目标路径应为文件夹而不是文件")
        filepaths = []
        datadict = {}
        listdir = filenames or os.listdir(path)
        for filename in listdir:
            filepath = path + "\\" + filename 
            if os.path.exists(filepath) and os.path.isfile(filepath) and self.__checkFileExtension(filepath):
                filepaths.append(filepath)
                wb = self.__getdata_to_workbook(filepath, area, sheets)
                datadict[filename] = wb
            else: continue
        return datadict

    # 获取文件目录下可选工作簿对象的可选工作表 返回dict {filename_workSheetName: ws}
    def __getSheet_in_folder(self, path, filenames=[], sheets=[], read_only=False, data_only=False)-> dict:
        """
        params:
        path: str 文件夹路径
        filename: list 工作簿名称列表
        sheets: list 工作表名称列表
        """
        if os.path.isfile(path): raise Exception(f"{path}目标路径应为文件夹而不是文件")
        listdir = filenames or os.listdir(path)
        sheet_map_dict = {}
        for filename in listdir:
            filepath = path + "\\" + filename 
            if os.path.exists(filepath) and os.path.isfile(filepath) and self.__checkFileExtension(filepath):
                wb = opx.load_workbook(filepath, read_only, data_only)
                if sheets: sheetlist = [wb[sheetname] for sheetname in sheets] 
                else:sheetlist = wb.worksheets
                for sheet in sheetlist:
                    map_name = f"{filename}_{sheet.title}"
                    sheet_map_dict[map_name] = sheet
        return sheet_map_dict

    # 以目标工作簿为模板创建多工作簿并根据cellMap写入数据
    def __createSheetsOnTemplate_to_workbook(self, path, cellMap)-> dict:
        """
        params:
        path: str
        cellMap: dict 二维字典 {workbookname1: {cell1: value1, cell2: value2...}, workbookname2: {cell1: value1, }...}
        """
        if not os.path.isfile(path): raise Exception("路径不是个文件")
        if not isinstance(cellMap, dict): raise Exception("cellMap参数类型错误")
        self.__checkFileExtension(path)
        self.workbook = opx.load_workbook(path)
        tp_wb = self.workbook
        tp_ws = tp_wb.active
        new_wbdict = {}
        for wb_map, cell_map in cellMap.items():
            new_wb = opx.Workbook()
            new_ws = new_wb.active
            new_filename = wb_map
            
            self.__copySheet(tp_ws, new_ws)
            for cell_name, value in cell_map.items():
                new_ws[cell_name].value = value
                    
            new_wbdict[new_filename] = new_wb
        return new_wbdict
    
    # 合并文件夹内可选工作簿可选工作表下的数据以工作表形式储存到新工作簿
    def __mergeWorkbooks_to_workbook(self, path, filenames=[], sheets=[])-> opx.Workbook:
        new_wb = opx.Workbook()
        sheet_map_dict = self.__getSheet_in_folder(path, filenames, sheets, read_only=True, data_only=True)
        ws_index = 1
        for map_name, ws in sheet_map_dict.items():
            new_ws = new_wb.create_sheet(f"{ws_index}_{ws.title}")
            self.__mergeSheet(ws, new_ws)
            ws_index += 1
        return new_wb
    
    def getdata_saveto_workbook(self, path, area={}, sheets=[])-> None:
        wb = self.__getdata_to_workbook(path, area, sheets)
        self.__save_file(wb, "newfile", "getdata_saveto_workbook")

    def getdata_saveto_workbooks(self, path, area={}, filenames=[], sheets=[])-> None:
        datadict = self.__getdata_for_workbooks(path, area, filenames, sheets)
        self.__save_files_in_folder(datadict, "getdata_saveto_workbooks")
        
    def splitSheets_saveto_workbooks(self, path, sheets=[])-> None:
        new_wbdict = self.__splitSheets_to_workbooks(path, sheets)
        self.__save_files_in_folder(new_wbdict, "forSplitSheets")
    
    def createSheetsOnTemplate_saveto_workbook(self, path, cellMap)-> None:
        new_wbdict = self.__createSheetsOnTemplate_to_workbook(path, cellMap)
        self.__save_files_in_folder(new_wbdict, "forTemplate")
                
    def mergeWorkbooks_saveto_newWorkbook(self, path, filenames=[], sheets=[])-> None:
        new_wb = self.__mergeWorkbooks_to_workbook(path, filenames, sheets)
        self.__save_file(new_wb, "newfile", "mergeWorkbooks_saveto_newWorkbook")
        
def getdata_saveto_workbook_test():
    path = r".\res\数据源_test1.xlsx"
    edm = ExcelDataModel()
    edm.getdata_saveto_workbook(path)

def splitSheets_saveto_workbooks_test():
    path = r".\res\数据源_test1.xlsx"
    edm = ExcelDataModel()
    edm.splitSheets_saveto_workbooks(path)

def createSheetsOnTemplate_saveto_workbook_test():
    path = r".\res\template.xlsx"
    """    
    cellmap = {
        文件名:{
            单元格: 值
        }
    }
    """
    cellmap = {
        1:{
            'a3':1, 
        }, 
        2:{
            'a3':2, 
        }, 
    }
    edm = ExcelDataModel()
    edm.createSheetsOnTemplate_saveto_workbook(path, cellmap)

def getdata_saveto_workbooks_test():
    path = r".\res"
    edm = ExcelDataModel()
    edm.getdata_saveto_workbooks(path)

def mergeWorkbooks_saveto_newWorkbook_test():
    path = r".\res"
    edm = ExcelDataModel()
    edm.mergeWorkbooks_saveto_newWorkbook(path)

if __name__ == '__main__':
    mergeWorkbooks_saveto_newWorkbook_test()