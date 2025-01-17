import openpyxl as opx
import os
import time

class ExcelDataModel:
    def __init__(self, workbook=opx.Workbook()) -> None:
        """
        ExcelDataModel 数据模型对象 数据获取 数据分析
        """
        self.workbook = workbook
        self.fileExtension = (".xlsx", ".xlrd")

    # 获取单工作簿多工作表数据
    def getdata(self, path, area:dict, sheets=[]) -> dict:
        """
        params: 
        path: str 文件路径
        area: dict 一维字典
        sheets: list 工作表名称列表
        """
        if not os.path.isfile(path): raise Exception(f"{path}不是一个文件")
        if not isinstance(area, dict): raise Exception("area参数类型应为一维dict对象")
        if not isinstance(sheets, list): raise Exception("sheets参数类型应为一维list对象")
        workbook = opx.load_workbook(path, read_only=True, data_only=True)
        if sheets: thissheets = [workbook[sheetname] for sheetname in sheets]
        else: thissheets = workbook.worksheets 
        data = {}
        for sheet in thissheets:
            sheetData = {}
            sheetname = sheet.title
            for map in area:
                cellindex = 1
                cellData = {}
                cells = sheet[area[map]]
                for celltuple in cells:
                    for cell in celltuple:
                        if not (isinstance(cell, opx.cell.read_only.ReadOnlyCell) or isinstance(cell, opx.cell.cell.Cell)): continue
                        celldatakey = map + "_" + str(cellindex)
                        cellData[celldatakey] = cell.value
                        cellindex += 1
                sheetData[map] = cellData
            data[sheetname] = sheetData
        return data

    # 获取多工作簿多工作表数据 后续调整异步
    def getdata_for_workbooks(self, path:str, area:dict, filenames=[], sheets=[]) -> list: 
        """
        params:
        path: str 文件夹路径
        area: dict 一维字典
        filename: list 工作簿名称列表
        sheets: list 工作表名称列表
        """
        if not isinstance(area, dict): raise Exception("area参数类型应为一维dict对象")
        if not isinstance(filenames, list): raise Exception("filenames参数类型应为一维list对象")
        if not isinstance(sheets, list): raise Exception("sheets参数类型应为一维list对象")
        if os.path.isfile(path): raise Exception("目标路径应为文件夹而不是文件")
        filepaths = []
        datalist = []
        listdir = filenames or os.listdir(path)
        for filename in listdir:
            filepath = path + "\\" + filename 
            if os.path.exists(filepath) and os.path.isfile(filepath):
                filepaths.append(filepath)
                data = self.getdata(filepath, area, sheets)
                datalist.append(data)
            else: continue
        return datalist

    # 以newfile.xlsx保存在本地
    def save_to_Excel(self, datalist) -> None:
        """
        params:
        datalist: list
        """
        if not isinstance(datalist, list): datalist = [datalist]
        wb = self.workbook
        ws = wb.active
        ws.append([])
        def iter_data(dictionary, path=[]):
            """
            深度优先遍历字典的函数
            :param dictionary: 要遍历的字典
            :param path: 存储当前路径，默认为空列表
            """
            for key, value in dictionary.items():
                # 将当前键添加到路径中
                new_path = path + [key]
                if isinstance(value, dict):
                    # 如果值是字典，递归调用 iter_data 函数
                    iter_data(value, new_path)
                else:
                    # 如果值不是字典，打印当前路径和对应的值
                    # print(f"Path: {new_path}, Value: {value}")
                    row = new_path + [value]
                    # datalist.append(row)
                    ws.append(row)
        for data in datalist: iter_data(data)
        wb.save(r".\newfile.xlsx")
    
    # 工作簿内多工作表数据拆分成多工作簿
    def splitsheets_saveto_workbooks(self, path, startindex=1, endindex=None):
        """
        params:
        path: str 目标文件路径
        startindex: int 起始索引
        endindex: int 结束索引
        """
        if not os.path.isfile(path): raise Exception("路径不是个文件")
        self.workbook = opx.load_workbook(path)
        wb = self.workbook
        worksheets = wb.worksheets
        endindex = endindex or len(worksheets)
        if startindex >= endindex: raise Exception("startindex参数不能大于endindex")
        worksheets = worksheets[(startindex - 1):endindex]
        filename = os.path.basename(path)
        timestamp = time.strftime("%Y_%m_%d_%H%M%S", time.localtime())
        newfolder = r".\newfiles" + timestamp
        os.mkdir(newfolder)
        count = 1
        for sheet in worksheets:
            newwb = opx.Workbook()
            newsheetname = sheet.title
            newsheet = newwb.active
            newsheet.title = newsheetname
            # 1. 数据先加进去
            for row in sheet.iter_rows():
                for cell in row:
                    rowindex = cell.row
                    colindex = cell.column
                    value = cell.value
                    newsheet.cell(rowindex, colindex, value)

            # 2. 处理合并单元格
            sheet_merged_cells = sheet.merged_cells.ranges
            if sheet_merged_cells:
                for merged_cell in sheet_merged_cells:
                    newsheet.merge_cells(str(merged_cell))
            
            # 3. 处理样式 有点麻烦 先放放吧

            newfilename =  f"{count}_" + newsheetname + filename
            newwb.save(f"./{newfolder}/{newfilename}")
            count += 1
 
def getdata_for_workbooks_test():
    path = r".\res"
    # 以单维字典 { 字段名:区域 } 的格式声明区域
    area = {
        "项":"d5:g7",
    }
    edm = ExcelDataModel()
    data = edm.getdata_for_workbooks(path, area)
    edm.save_to_Excel(data)

def getdata_test():
    path = r".\res\数据源_test1.xlsx"
    # 以单维字典 { 字段名:区域 } 的格式声明区域
    area = {
        "项":"d5:g7",
    }
    edm = ExcelDataModel()
    data = edm.getdata(path, area)
    edm.save_to_Excel(data)

def splitsheets_saveto_workbooks_test():
    path = r".\res\数据源_test1.xlsx"
    # 以单维字典 { 字段名:区域 } 的格式声明区域
    edm = ExcelDataModel()
    edm.split_sheets_to_workbooks(path)

if __name__ == '__main__':
    splitsheets_saveto_workbooks_test()